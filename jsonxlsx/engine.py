"""Main JSONL-to-XLSX engine: load, reduce, render."""

import argparse
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from jsonxlsx.reduce import init_schema, reduce_log
from jsonxlsx.render import render_preserve, render_sheet_data_replace


STRATEGY_DISPATCH = {
    "preserve": render_preserve,
    "data_replace": render_sheet_data_replace,
}


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    """Load JSONL file line-by-line, skip blank lines."""
    if not path.exists():
        return []
    out: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if s:
            out.append(json.loads(s))
    return out


def main() -> None:
    """
    Main entry point: masters + edges + config → XLSX.

    Steps:
    1. Load and reduce schema from config/schema.jsonl
    2. Load and reduce masters from masters/*.jsonl
    3. Load and reduce edges from edges/*.jsonl
    4. Load sheet config from config/sheets.jsonl
    5. Load template XLSX
    6. For each sheet config, apply strategy (preserve or data_replace)
    7. Write output XLSX
    """
    parser = argparse.ArgumentParser(
        description="masters + edges + config → xlsx (pure renderer)"
    )
    parser.add_argument(
        "--template", required=True, help="Reference XLSX template (format/style source)"
    )
    parser.add_argument("--output", required=True, help="Output XLSX path")
    parser.add_argument("--masters", required=True, help="Directory containing masters/*.jsonl")
    parser.add_argument("--edges", required=True, help="Directory containing edges/*.jsonl")
    parser.add_argument("--config", required=True, help="Directory containing config/*.jsonl")
    args = parser.parse_args()

    template_path = Path(args.template)
    output_path = Path(args.output)
    masters_dir = Path(args.masters)
    edges_dir = Path(args.edges)
    config_dir = Path(args.config)

    if not template_path.exists():
        print(f"ERROR: template not found: {template_path}", file=sys.stderr)
        sys.exit(1)

    init_schema(reduce_log(load_jsonl(config_dir / "schema.jsonl")))

    masters = {
        p.stem: reduce_log(load_jsonl(p)) for p in sorted(masters_dir.glob("*.jsonl"))
    }
    edges = {p.stem: reduce_log(load_jsonl(p)) for p in sorted(edges_dir.glob("*.jsonl"))}
    sheets_cfg = reduce_log(load_jsonl(config_dir / "sheets.jsonl"))

    print(f"[load] template: {template_path}")
    print(f"[load] schema : {len(masters.get('_schema_types', []))} types")
    print(f"[load] masters: {sorted(masters.keys())}")
    print(f"[load] edges  : {sorted(edges.keys())}")
    print(f"[load] sheets : {len(sheets_cfg)}")

    type_counts: Counter = Counter()
    generic_samples: list[tuple[str, list[str]]] = []
    for src, recs in (
        [(f"masters/{k}", v) for k, v in masters.items()]
        + [(f"edges/{k}", v) for k, v in edges.items()]
    ):
        for r in recs:
            from jsonxlsx.reduce import _record_key
            prefix = _record_key(r)[0]
            type_counts[prefix] += 1
            if prefix == "generic" and len(generic_samples) < 5:
                generic_samples.append(
                    (src, sorted(k for k in r.keys() if not k.startswith("_")))
                )

    print(f"[schema] key distribution: {dict(sorted(type_counts.items()))}")
    if type_counts.get("generic", 0):
        print(f"⚠ {type_counts['generic']} record(s) → generic fallback:")
        for src, keys in generic_samples:
            print(f"    {src}: fields={keys}")

    wb = load_workbook(str(template_path))

    for cfg in sheets_cfg:
        sheet_name = cfg["sheet"]
        strategy = cfg.get("strategy", "data_replace")
        if sheet_name not in wb.sheetnames:
            print(f"  [skip] {sheet_name}: not in template")
            continue
        handler = STRATEGY_DISPATCH.get(strategy)
        if handler is None:
            raise ValueError(f"unknown strategy: {strategy} (sheet: {sheet_name})")
        ws = wb[sheet_name]
        n = (
            len(masters.get(cfg.get("source", ""), []))
            if strategy == "data_replace"
            else 0
        )
        print(f"  [{strategy}] {sheet_name} ({n} records)")
        handler(ws, cfg, masters, edges)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"\n[OK] wrote: {output_path}")


if __name__ == "__main__":
    main()
